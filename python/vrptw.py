import openpyxl
from copy import *
import time
start_time = time.time()

# variables
clientNum = 50  # change this when necessary
nurseNum = 3  # nurse 0-2
dayNum = 10  # day 0-9
maxq_dif = 240
maxq = 480
E = 0
L = 480
M = 5000

# loading from excel
# change this when necessary
wb = openpyxl.load_workbook(filename='50c_3n_5.xlsx')
wscost = wb['cost data']
wsdemand = wb['demand data']
wsstart = wb['start']
wsend = wb['end']

cost = []  # cost[i][j]
for row in range(2, clientNum+4):
    temp = []
    for col in range(2, clientNum+4):
        temp.append(wscost.cell(row, col).value)
    cost.append(temp)

demand = []  # demand[i]
for col in range(2, clientNum+4):
    demand.append(wsdemand.cell(2, col).value)

start = []  # start[i][day]
for row in range(2, clientNum+4):
    temp = []
    for col in range(2, dayNum+2):
        temp.append(wsstart.cell(row, col).value)
    start.append(temp)

end = []  # end[i][day]
for row in range(2, clientNum+4):
    temp = []
    for col in range(2, dayNum+2):
        temp.append(wsend.cell(row, col).value)
    end.append(temp)

# service start time[day][nurse][vertix], waiting time[day][nurse][vertix], leaving time[day][nurse][vertix]
waiting = [[[0 for x in range(clientNum+2)]
            for x in range(nurseNum)] for x in range(dayNum)]
servicestart = [[[0 for x in range(clientNum+2)]
                 for x in range(nurseNum)] for x in range(dayNum)]
serviceend = [[[0 for x in range(clientNum+2)]
               for x in range(nurseNum)] for x in range(dayNum)]


# find out the current max saving pair
def maxsavings(savings, usedconnections, unableconnections):
    maxi = 0
    maxj = 0
    max = 0
    for i in range(1, clientNum+1):
        for j in range(i+1, clientNum+1):
            if savings[i][j] > max and {i, j} not in usedconnections and {i, j} not in unableconnections:
                maxi = i
                maxj = j
                max = savings[i][j]
    return maxi, maxj  # returns a tuple


# start, end, and waiting time for a certain route
def routetimeings(route, day):
    tempservicestart = [0]*(clientNum+2)
    tempwaiting = [0]*(clientNum+2)
    tempserviceend = [0]*(clientNum+2)

    for index, node in enumerate(route):
        if node == 0:
            pass
        else:
            arrivaltime = tempserviceend[route[index-1]
                                         ] + cost[route[index-1]][node]
            if arrivaltime < start[node][day]:
                tempwaiting[node] = start[node][day] - arrivaltime
                tempservicestart[node] = start[node][day]
                tempserviceend[node] = tempservicestart[node] + demand[node]
            else:
                tempwaiting[node] = 0
                tempservicestart[node] = arrivaltime
                tempserviceend[node] = tempservicestart[node] + demand[node]

    return tempwaiting, tempservicestart, tempserviceend


# check whether is valid route (time window and capacity)
def isvalidroute(route, day):
    nodewaiting, nodestart, nodeend = routetimeings(route, day)
    # returns 0 if not in time window
    for index, endtime in enumerate(nodeend):
        if endtime > end[index][day]:
            return 0

    # returns 0 if exceeds capacity
    servicetime = 0
    for node in route:
        servicetime = servicetime + demand[node]
    if servicetime > maxq:
        return 0

    return 1


def test():
    print('constructing savings...')
    savings = [[0 for x in range(clientNum+2)] for x in range(clientNum+2)]
    for x in range(1, clientNum+2):
        for y in range(1, clientNum+2):
            if x == y:
                continue
            savings[x][y] = cost[0][x]+cost[0][y]-cost[x][y]

    usedconnections = []
    infeasiable = []
    route = [0, 33]
    day = 0
    i, j = maxsavings(savings, usedconnections, infeasiable)
    print(f'{i}, {j}')

    infeasiable.append({i, j})
    i, j = maxsavings(savings, usedconnections, infeasiable)
    print(f'{i}, {j}')

    route = [0, 26, 25, 33]
    if isvalidroute(route, day):
        print('route is able')
    else:
        print('invalid route')


def main():
    # construct savings matrix
    # savings[vertix][vertix]
    print('constructing savings...')
    savings = [[0 for x in range(clientNum+2)] for x in range(clientNum+2)]
    for x in range(1, clientNum+2):
        for y in range(1, clientNum+2):
            if x == y:
                continue
            savings[x][y] = cost[0][x]+cost[0][y]-cost[x][y]

    # construcing routes
    # routes[day][nurse]
    # used connections = [{}, {}] --> sets in a list
    usedconnections = []
    visitedvertex = set()
    allnodeset = set()
    for node in range(1, clientNum+1):
        allnodeset.add(node)
    routes = []
    for day in range(dayNum):
        dayroute = []
        for nurse in range(nurseNum):
            nurseroute = [0, clientNum+1]
            dayroute.append(nurseroute)
        routes.append(dayroute)

    # constructing routes for each day for each nurse
    for day in range(dayNum):
        print(f'construcing routes for day {day}...')
        infeasiableconnections = []  # = [{}, {}] --> sets in a list
        while 1:
            # finding out max savings
            maxi, maxj = maxsavings(
                savings, usedconnections, infeasiableconnections)

            # end of day
            if maxi == 0 and maxj == 0:
                print(f'end of day {day}')
                print(f'{routes[day]}')
                print(f'total used connections {usedconnections}')
                print(f'total visited {visitedvertex}')
                break

            print(f'trying ({maxi}, {maxj})')
            for nurse in range(nurseNum):
                temproute = deepcopy(routes[day][nurse])

                # if route is empty
                if temproute == [0, clientNum+1]:
                    if maxi in visitedvertex or maxj in visitedvertex:  # move to next nurse
                        print(f'nurse {nurse} infeasiable...')
                        continue
                    else:  # insert nodes and check feasiability
                        if end[maxi][day] <= end[maxj][day]:
                            temproute = [0, maxi, maxj, clientNum+1]
                            if isvalidroute(temproute, day):
                                print(
                                    f'route is valid, current route {temproute}...')
                                routes[day][nurse] = deepcopy(temproute)
                                usedconnections.append({maxi, maxj})
                                visitedvertex.add(maxi)
                                visitedvertex.add(maxj)
                                break

                            temproute = [0, maxj, maxi, clientNum+1]
                            if isvalidroute(temproute, day):
                                print(
                                    f'route is valid, current route {temproute}...')
                                routes[day][nurse] = deepcopy(temproute)
                                usedconnections.append({maxi, maxj})
                                visitedvertex.add(maxi)
                                visitedvertex.add(maxj)
                                break

                            print(f'nurse {nurse} is infeasiable')
                            print(
                                f'nurse {nurse} current route: {routes[day][nurse]}')
                            continue

                        else:
                            temproute = [0, maxj, maxi, clientNum+1]
                            if isvalidroute(temproute, day):
                                print(
                                    f'route is valid, current route {temproute}...')
                                routes[day][nurse] = deepcopy(temproute)
                                usedconnections.append({maxi, maxj})
                                visitedvertex.add(maxi)
                                visitedvertex.add(maxj)
                                break

                            temproute = [0, maxi, maxj, clientNum+1]
                            if isvalidroute(temproute, day):
                                print(
                                    f'route is valid, current route {temproute}...')
                                routes[day][nurse] = deepcopy(temproute)
                                usedconnections.append({maxi, maxj})
                                visitedvertex.add(maxi)
                                visitedvertex.add(maxj)
                                break

                            print(f'nurse {nurse} is infeasiable')
                            print(
                                f'nurse {nurse} current route: {routes[day][nurse]}')
                            continue

                # if route is not empty
                else:
                    if maxi == temproute[1] and maxj not in visitedvertex:
                        temproute.insert(1, maxj)
                        if isvalidroute(temproute, day):
                            print(
                                f'route is valid, current route {temproute}...')
                            routes[day][nurse] = deepcopy(temproute)
                            usedconnections.append({maxi, maxj})
                            visitedvertex.add(maxj)
                            break

                    elif maxi == temproute[len(temproute)-2] and maxj not in visitedvertex:
                        temproute.insert(len(temproute)-1, maxj)
                        if isvalidroute(temproute, day):
                            print(
                                f'route is valid, current route {temproute}...')
                            routes[day][nurse] = deepcopy(temproute)
                            usedconnections.append({maxi, maxj})
                            visitedvertex.add(maxj)
                            break

                    elif maxj == temproute[1] and maxi not in visitedvertex:
                        temproute.insert(1, maxi)
                        if isvalidroute(temproute, day):
                            print(
                                f'route is valid, current route {temproute}...')
                            routes[day][nurse] = deepcopy(temproute)
                            usedconnections.append({maxi, maxj})
                            visitedvertex.add(maxi)
                            break

                    elif maxj == temproute[len(temproute)-2] and maxi not in visitedvertex:
                        temproute.insert(len(temproute)-1, maxi)
                        if isvalidroute(temproute, day):
                            print(
                                f'route is valid, current route {temproute}...')
                            routes[day][nurse] = deepcopy(temproute)
                            usedconnections.append({maxi, maxj})
                            visitedvertex.add(maxi)
                            break

                    print(f'nurse {nurse} is infeasiable')
                    print(f'nurse {nurse} current route: {routes[day][nurse]}')
                    continue

            # checking whether the pair is feasiable
            newmaxi, newmaxj = maxsavings(
                savings, usedconnections, infeasiableconnections)
            if maxi == newmaxi and maxj == newmaxj:
                print(f'{maxi, maxj} is infeasiable')
                infeasiableconnections.append({maxi, maxj})

        # check all nodes visited
        if visitedvertex == allnodeset:
            print('all routes constructed')
            for day in range(dayNum):
                print(f'day {day}:')
                print(routes[day])
            print(f'total used connections {usedconnections}')
            print(f'total visited {visitedvertex}')
            break

    # if not all nodes inserted
    if visitedvertex != allnodeset:
        remainingnodes = allnodeset.difference(visitedvertex)
        print(f'remaining {remainingnodes}')
        for day in range(dayNum):
            if len(remainingnodes) == 0:
                break

            for nurse in range(nurseNum):
                if len(remainingnodes) == 0:
                    break

                if routes[day][nurse] == [0, clientNum+1]:
                    node = remainingnodes.pop()
                    temproute = [0, node, clientNum+1]

                    print(f'trying {temproute}...')
                    if isvalidroute(temproute, day):
                        routes[day][nurse] = deepcopy(temproute)
                        visitedvertex.add(node)
                        print(
                            f'day {day} nurse {nurse}: {routes[day][nurse]}')
                        print(f'total visited {visitedvertex}')
                        print(f'remaining {remainingnodes}')
                        continue
                    else:
                        remainingnodes.add(node)
                        print(f'node {node} unable')

    if visitedvertex == allnodeset:
        print('all routes constructed')
        for day in range(dayNum):
            print(f'day {day}:')
            print(routes[day])
        print(f'total used connections {usedconnections}')
        print(f'total visited {visitedvertex}')

    # calculate travel time
    nursetraveltime = [0]*nurseNum
    totaltraveltime = 0
    for nurse in range(nurseNum):
        for day in range(dayNum):
            for index in range(len(routes[day][nurse])-1):
                nursetraveltime[nurse] = nursetraveltime[nurse] + \
                    cost[routes[day][nurse][index]][routes[day][nurse][index+1]]
        print(f'nurse {nurse}: {nursetraveltime[nurse]} minutes')
        totaltraveltime = totaltraveltime + nursetraveltime[nurse]
    print(f'total travel time: {totaltraveltime} minutes')


main()
print("--- %s seconds ---" % (time.time() - start_time))
